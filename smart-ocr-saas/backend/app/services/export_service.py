"""
Export service for generating output files.
"""
import io
import json
import csv
from typing import Tuple


class ExportService:
    """Service for exporting OCR results."""

    async def export(self, task_id: str, format: str) -> Tuple[bytes, str, str]:
        """
        Export OCR results in specified format.

        Args:
            task_id: Task ID
            format: Export format (json, csv, xlsx)

        Returns:
            Tuple of (content bytes, media type, filename)
        """
        from app.repositories.ocr_result_repository import OCRResultRepository
        from app.repositories.ocr_task_repository import OCRTaskRepository

        result_repo = OCRResultRepository()
        task_repo = OCRTaskRepository()

        task = await task_repo.find_by_id(task_id)
        results = await result_repo.find_by_task(task_id)

        base_filename = task["original_filename"].rsplit(".", 1)[0]

        if format == "json":
            content = self._export_json(results)
            return content, "application/json", f"{base_filename}_ocr.json"

        elif format == "csv":
            content = self._export_csv(results)
            return content, "text/csv; charset=utf-8-sig", f"{base_filename}_ocr.csv"

        elif format == "xlsx":
            content = self._export_xlsx(results, base_filename)
            return (
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"{base_filename}_ocr.xlsx",
            )

        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_json(self, results: list) -> bytes:
        """
        Export results as JSON.

        Args:
            results: List of OCR result dicts

        Returns:
            JSON bytes
        """
        export_data = []

        for result in results:
            export_data.append({
                "page_number": result.get("page_number", 1),
                "extracted_text": result.get("extracted_text", ""),
                "structured_data": result.get("structured_data", {}),
                "confidence": result.get("confidence", 0),
            })

        return json.dumps(
            export_data,
            ensure_ascii=False,
            indent=2,
        ).encode("utf-8")

    def _export_csv(self, results: list) -> bytes:
        """
        Export results as CSV.

        Args:
            results: List of OCR result dicts

        Returns:
            CSV bytes with UTF-8 BOM
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(["頁碼", "辨識文字", "文件類型", "信心度"])

        # Data rows
        for result in results:
            structured = result.get("structured_data", {})
            doc_type = structured.get("type", "") if structured else ""

            writer.writerow([
                result.get("page_number", 1),
                result.get("extracted_text", ""),
                doc_type,
                result.get("confidence", 0),
            ])

        # Return with UTF-8 BOM for Excel compatibility
        return output.getvalue().encode("utf-8-sig")

    def _export_xlsx(self, results: list, base_filename: str) -> bytes:
        """
        Export results as Excel XLSX.

        Args:
            results: List of OCR result dicts
            base_filename: Base filename for the sheet

        Returns:
            XLSX bytes
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()

        # Main sheet with structured fields (primary view)
        ws = wb.active
        ws.title = "辨識結果"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Target fields in specific order
        target_fields = [
            "銀行名稱", "戶名", "身分證", "存款帳號",
            "連絡電話", "聯絡地址", "電費", "水費", "電信費"
        ]

        # Headers: 頁碼 + target fields
        headers = ["頁碼"] + target_fields
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows - one row per page/result
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result.get("page_number", 1)).border = thin_border

            # Extract fields into a dict for easy lookup
            structured = result.get("structured_data", {})
            fields = structured.get("fields", []) if structured else []
            fields_dict = {f.get("key", ""): f.get("value", "") for f in fields}

            # Fill in each target field column
            for col_idx, field_name in enumerate(target_fields, 2):
                value = fields_dict.get(field_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Column widths
        ws.column_dimensions["A"].width = 6   # 頁碼
        ws.column_dimensions["B"].width = 15  # 銀行名稱
        ws.column_dimensions["C"].width = 12  # 戶名
        ws.column_dimensions["D"].width = 14  # 身分證
        ws.column_dimensions["E"].width = 18  # 存款帳號
        ws.column_dimensions["F"].width = 20  # 連絡電話
        ws.column_dimensions["G"].width = 35  # 聯絡地址
        ws.column_dimensions["H"].width = 14  # 電費
        ws.column_dimensions["I"].width = 14  # 水費
        ws.column_dimensions["J"].width = 14  # 電信費

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add raw text sheet
        self._add_raw_text_sheet(wb, results)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _add_raw_text_sheet(self, wb, results: list):
        """Add sheet with raw extracted text."""
        ws = wb.create_sheet("原始文字")

        ws.append(["頁碼", "辨識文字", "信心度"])

        for result in results:
            ws.append([
                result.get("page_number", 1),
                result.get("extracted_text", ""),
                result.get("confidence", 0),
            ])

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 100
        ws.column_dimensions["C"].width = 10

    def _add_fields_sheet(self, wb, results: list):
        """Add sheet with extracted fields."""
        ws = wb.create_sheet("提取欄位")

        # Headers
        ws.append(["頁碼", "欄位名稱", "欄位值", "信心度"])

        for result in results:
            structured = result.get("structured_data", {})
            fields = structured.get("fields", []) if structured else []

            for field in fields:
                ws.append([
                    result.get("page_number", 1),
                    field.get("key", ""),
                    field.get("value", ""),
                    field.get("confidence", 0),
                ])

    def _add_tables_sheet(self, wb, results: list):
        """Add sheet with extracted tables."""
        ws = wb.create_sheet("提取表格")

        row_num = 1

        for result in results:
            structured = result.get("structured_data", {})
            tables = structured.get("tables", []) if structured else []

            for table_idx, table in enumerate(tables):
                # Table header
                ws.cell(row=row_num, column=1, value=f"頁 {result.get('page_number', 1)} - 表格 {table_idx + 1}")
                row_num += 1

                # Table headers
                headers = table.get("headers", [])
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col, value=header)
                row_num += 1

                # Table rows
                for table_row in table.get("rows", []):
                    for col, cell_value in enumerate(table_row, 1):
                        ws.cell(row=row_num, column=col, value=cell_value)
                    row_num += 1

                row_num += 1  # Empty row between tables
