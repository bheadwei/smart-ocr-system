"""
OCR Engine Module

Core OCR processing engine using PaddleOCR for text recognition.
"""

import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from .config import OCRConfig

logger = logging.getLogger(__name__)


def _setup_cuda_environment() -> None:
    """Setup CUDA/cuDNN environment variables for Windows."""
    try:
        import nvidia.cudnn
        import nvidia.cublas

        cudnn_path = Path(nvidia.cudnn.__file__).parent / "bin"
        cublas_path = Path(nvidia.cublas.__file__).parent / "bin"

        current_path = os.environ.get("PATH", "")
        new_paths = [str(cudnn_path), str(cublas_path)]

        for p in new_paths:
            if p not in current_path:
                os.environ["PATH"] = p + ";" + current_path
                current_path = os.environ["PATH"]

        logger.debug(f"Added CUDA paths: {new_paths}")
    except ImportError:
        logger.debug("nvidia-cudnn/cublas not installed, skipping CUDA setup")


class OCREngine:
    """
    High-performance OCR engine for Chinese text recognition.

    This engine wraps PaddleOCR and provides:
    - Single image processing
    - Batch processing for directories
    - Multiple export format support
    - Configurable recognition parameters

    Example:
        >>> config = OCRConfig.for_chinese()
        >>> engine = OCREngine(config)
        >>> result = engine.process_image("document.png")
        >>> print(result.text)
    """

    def __init__(self, config: Optional[OCRConfig] = None):
        """
        Initialize the OCR engine.

        Args:
            config: OCR configuration. If None, uses default Chinese config.
        """
        self.config = config or OCRConfig.for_chinese()
        self._ocr = None
        self._initialized = False

    def _ensure_initialized(self) -> None:
        """Lazily initialize PaddleOCR engine."""
        if self._initialized:
            return

        # Setup CUDA environment for GPU mode
        if self.config.use_gpu:
            _setup_cuda_environment()

        try:
            from paddleocr import PaddleOCR

            logger.info(
                f"Initializing PaddleOCR (GPU: {self.config.use_gpu}, "
                f"Lang: {self.config.lang})"
            )

            # PaddleOCR 2.x API
            self._ocr = PaddleOCR(
                use_angle_cls=self.config.use_angle_cls,
                lang=self.config.lang,
                use_gpu=self.config.use_gpu,
                show_log=self.config.show_log,
            )
            self._initialized = True
            logger.info("PaddleOCR engine initialized successfully")

        except ImportError:
            raise ImportError(
                "PaddleOCR is not installed. "
                "Please install it with: pip install paddlepaddle paddleocr"
            )

    def process_image(self, image_path: Union[str, Path]) -> "OCRResult":
        """
        Process a single image and extract text.

        Args:
            image_path: Path to the image file.

        Returns:
            OCRResult containing extracted text and metadata.

        Raises:
            FileNotFoundError: If image file doesn't exist.
            OCRError: If OCR processing fails.
        """
        self._ensure_initialized()

        image_path = Path(image_path)
        if not image_path.exists():
            raise FileNotFoundError(f"Image not found: {image_path}")

        logger.info(f"Processing image: {image_path}")

        # PaddleOCR 2.x API
        assert self._ocr is not None, "OCR engine not initialized"
        result = self._ocr.ocr(str(image_path), cls=self.config.use_angle_cls)

        return OCRResult.from_paddle_result(result, image_path)

    def process_directory(
        self,
        directory: Union[str, Path],
        extensions: Optional[List[str]] = None,
        recursive: bool = False,
    ) -> List["OCRResult"]:
        """
        Process all images in a directory.

        Args:
            directory: Path to the directory containing images.
            extensions: List of file extensions to process.
                       Default: ['.png', '.jpg', '.jpeg', '.bmp', '.tiff']
            recursive: Whether to search subdirectories.

        Returns:
            List of OCRResult objects.
        """
        directory = Path(directory)
        if not directory.is_dir():
            raise NotADirectoryError(f"Not a directory: {directory}")

        extensions = extensions or [".png", ".jpg", ".jpeg", ".bmp", ".tiff"]
        extensions = [
            ext.lower() if ext.startswith(".") else f".{ext.lower()}"
            for ext in extensions
        ]

        pattern = "**/*" if recursive else "*"
        image_files = [
            f
            for f in directory.glob(pattern)
            if f.is_file() and f.suffix.lower() in extensions
        ]

        logger.info(f"Found {len(image_files)} images to process")

        results = []
        for image_path in sorted(image_files):
            try:
                result = self.process_image(image_path)
                results.append(result)
            except Exception as e:
                logger.error(f"Failed to process {image_path}: {e}")

        return results

    def export_results(
        self,
        results: List["OCRResult"],
        output_path: Optional[Union[str, Path]] = None,
        format: str = "json",
    ) -> Path:
        """
        Export OCR results to file.

        Args:
            results: List of OCRResult objects.
            output_path: Output file path. If None, auto-generates.
            format: Export format ('txt', 'json', 'csv').

        Returns:
            Path to the exported file.
        """
        if format not in self.config.export_formats:
            raise ValueError(f"Format {format} not enabled in config")

        output_path = output_path or self.config.output_dir / f"ocr_results.{format}"
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if format == "json":
            self._export_json(results, output_path)
        elif format == "csv":
            self._export_csv(results, output_path)
        elif format == "xlsx":
            self._export_xlsx(results, output_path)
        else:
            self._export_txt(results, output_path)

        logger.info(f"Results exported to: {output_path}")
        return output_path

    def _export_json(self, results: List["OCRResult"], path: Path) -> None:
        """Export results as JSON."""
        import json

        data = [r.to_dict() for r in results]
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _export_csv(self, results: List["OCRResult"], path: Path) -> None:
        """Export results as CSV."""
        import csv

        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["file", "text", "confidence", "bbox"])
            for r in results:
                for line in r.lines:
                    writer.writerow(
                        [r.source_file, line.text, line.confidence, line.bbox]
                    )

    def _export_txt(self, results: List["OCRResult"], path: Path) -> None:
        """Export results as plain text."""
        with open(path, "w", encoding="utf-8") as f:
            for r in results:
                f.write(f"=== {r.source_file} ===\n")
                f.write(r.text)
                f.write("\n\n")

    def _export_xlsx(self, results: List["OCRResult"], path: Path) -> None:
        """Export results as Excel file (UTF-8 encoded)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "OCR Results"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Write headers
        headers = ["檔案", "行號", "文字內容", "信心分數", "座標"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        row_num = 2
        for r in results:
            for line_idx, line in enumerate(r.lines, 1):
                ws.cell(row=row_num, column=1, value=str(r.source_file)).border = thin_border
                ws.cell(row=row_num, column=2, value=line_idx).border = thin_border
                ws.cell(row=row_num, column=3, value=line.text).border = thin_border
                ws.cell(row=row_num, column=4, value=f"{line.confidence:.2%}").border = thin_border
                ws.cell(row=row_num, column=5, value=str(line.bbox)).border = thin_border
                row_num += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30  # 檔案
        ws.column_dimensions["B"].width = 8   # 行號
        ws.column_dimensions["C"].width = 60  # 文字內容
        ws.column_dimensions["D"].width = 12  # 信心分數
        ws.column_dimensions["E"].width = 40  # 座標

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(path)


class OCRLine:
    """Represents a single line of recognized text."""

    def __init__(self, text: str, confidence: float, bbox: List[List[float]]):
        self.text = text
        self.confidence = confidence
        self.bbox = bbox

    def to_dict(self) -> Dict:
        return {"text": self.text, "confidence": self.confidence, "bbox": self.bbox}


class OCRResult:
    """Container for OCR processing results."""

    def __init__(self, source_file: Path, lines: List[OCRLine]):
        self.source_file = source_file
        self.lines = lines

    @property
    def text(self) -> str:
        """Get full text as a single string."""
        return "\n".join(line.text for line in self.lines)

    @property
    def average_confidence(self) -> float:
        """Calculate average confidence score."""
        if not self.lines:
            return 0.0
        return sum(line.confidence for line in self.lines) / len(self.lines)

    @classmethod
    def from_paddle_result(cls, result: Any, source_file: Path) -> "OCRResult":
        """Create OCRResult from PaddleOCR 2.x output.

        PaddleOCR 2.x format:
        [
          [  # page results
            [bbox, (text, confidence)],
            [bbox, (text, confidence)],
            ...
          ]
        ]
        """
        lines = []
        if result and result[0]:
            for item in result[0]:
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    bbox = item[0] if isinstance(item[0], list) else []
                    if isinstance(item[1], (list, tuple)) and len(item[1]) >= 2:
                        text = str(item[1][0])
                        confidence = float(item[1][1])
                        lines.append(OCRLine(text, confidence, bbox))
        return cls(source_file, lines)

    def to_dict(self) -> Dict:
        """Convert to dictionary for serialization."""
        return {
            "source_file": str(self.source_file),
            "text": self.text,
            "average_confidence": self.average_confidence,
            "lines": [line.to_dict() for line in self.lines],
        }
